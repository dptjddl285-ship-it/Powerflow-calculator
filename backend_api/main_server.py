# 파일명: main_server.py
from fastapi import FastAPI, Request, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from ultralytics import YOLO
import pandas as pd
import io
import os
from pathlib import Path

import core.env_loader

try:
    from agent_tools.vision_tools import configure_review_tools
    from core.power_logic import construct_y_bus
    from core.adaptive_vision_pipeline import analyze_circuit_image_adaptive
    from core.bus_number_linker import (
        link_and_validate_bus_numbers,
        propagate_bus_numbers_to_devices,
        synchronize_node_and_line_ids,
    )
    from review.api import router as review_router
    from review.staged_api import (
        configure_staged_review_model,
        router as staged_review_router,
    )
    from review.store import review_store
    from review.vision_adapter import build_graph_document
except ImportError as e:
    print(f"❌ 에러: 파일을 찾을 수 없습니다. {e}")

app = FastAPI()
app.include_router(review_router)
app.include_router(staged_review_router)
yolo_model = None
configure_staged_review_model(lambda: yolo_model)


def _run_review_vision(image_bytes: bytes):
    if yolo_model is None:
        raise RuntimeError("Vision model is not loaded")
    return analyze_circuit_image_adaptive(image_bytes, yolo_model)


configure_review_tools(_run_review_vision)

@app.on_event("startup")
async def startup_event():
    global yolo_model
    print("\n" + "★"*60)
    print(" 🚀 [PowerLens] 통합 서버 가동 시작!")
    try:
        # 2026_07_30_coslr.pt: cos_lr + 런타임 증강 켜서 재학습한 최신 모델.
        # 24bus/IEEE24bus 변압기까지 잡힘(이전 aug모델은 0개). vision_logic imgsz=960과 세트.
        # The local checkpoint remains the safe default.  A friend/all-class
        # checkpoint can be selected for the same hybrid pipeline without
        # changing code: POWERLENS_YOLO_MODEL=<path-to-best.pt>.
        default_model_path = (
            Path(__file__).resolve().parent
            / "models"
            / "2026_07_30_coslr.pt"
        )
        model_path = os.environ.get("POWERLENS_YOLO_MODEL", str(default_model_path))
        yolo_model = YOLO(model_path)
        print(" ✅ AI 모델(YOLO) 로딩 완료!")
    except Exception as e:
        print(f" ❌ 모델 로딩 실패: {e}")
    print("★"*60 + "\n")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/health")
async def health_check():
    return {
        "status": "ok",
        "model_loaded": yolo_model is not None,
    }

# [API 1] 사진 분석 
@app.post("/analyze_image")
async def process_image(file: UploadFile = File(...)):
    print("\n📸 [사진 수신] 분석 요청이 들어왔습니다.")
    try:
        image_bytes = await file.read()
        result_data = analyze_circuit_image_adaptive(image_bytes, yolo_model)
        
        # Link & Validate Bus Numbers (Runs after all objects are masked and rescued)
        if "nodes" in result_data and result_data["nodes"]:
            result_data["nodes"], bus_report = link_and_validate_bus_numbers(image_bytes, result_data["nodes"])
            result_data["nodes"] = propagate_bus_numbers_to_devices(result_data["nodes"], result_data.get("lines", []))
            result_data["nodes"], result_data["lines"] = synchronize_node_and_line_ids(result_data["nodes"], result_data.get("lines", []))
            result_data["bus_number_report"] = bus_report

        # Keep the legacy nodes/lines response for the current Flutter canvas,
        # and attach the versioned review contract for the new Agent workflow.
        graph_document = build_graph_document(
            result_data,
            image_bytes=image_bytes,
            filename=file.filename,
        )
        review_store.put(graph_document, overwrite=True)
        review_store.put_analysis_asset(
            graph_document.document_id,
            image_bytes=image_bytes,
            vision_result=result_data,
        )
        result_data["graph_document"] = graph_document.model_dump(mode="json")
        return {"status": "success", "data": result_data}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ==========================================
# 🎯 [API 2 - 엑셀 업로드 및 계통 파라미터 매핑]
# ==========================================
from core.excel_case_importer import ExcelCaseImporter

excel_importer = ExcelCaseImporter()


@app.post("/upload_excel")
async def upload_excel(file: UploadFile = File(...)):
    print(f"\n📂 [엑셀 수신] 업로드된 파일명: {file.filename}")
    try:
        contents = await file.read()
        parsed_case = excel_importer.parse_excel(contents)
        print(
            f"✅ 엑셀 파싱 성공! 슬랙 모선: #{parsed_case['slack_bus_number']}, "
            f"모선({parsed_case['total_buses']}개), 발전기({parsed_case['total_generators']}개), "
            f"선로({parsed_case['total_branches']}개)"
        )
        return {"status": "success", "data": parsed_case}
    except Exception as e:
        print(f"❌ 엑셀 처리 중 에러 발생: {e}")
        return {"status": "error", "message": str(e)}


@app.get("/load_default_excel")
async def load_default_excel():
    sample_path = Path(__file__).parent / "sample_cases" / "ac_case25.xlsx"
    if not sample_path.exists():
        return {"status": "error", "message": "기본 ac_case25 샘플 파일을 찾을 수 없습니다."}
    try:
        with open(sample_path, "rb") as f:
            contents = f.read()
        parsed_case = excel_importer.parse_excel(contents)
        print(f"✅ 기본 엑셀(ac_case25) 로드 성공! 슬랙 모선: #{parsed_case['slack_bus_number']}")
        return {"status": "success", "data": parsed_case}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.post("/apply_excel_to_elements")
async def apply_excel_to_elements(request: Request):
    try:
        body = await request.json()
        elements = body.get("elements", [])
        excel_data = body.get("excel_data", {})
        updated_elements, summary = excel_importer.apply_to_elements(elements, excel_data)
        return {
            "status": "success",
            "elements": updated_elements,
            "summary": summary
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}

# [API 3] 조류 계산 (AC Newton-Raphson Solver)
from core.power_flow_solver import PowerFlowSolver
import io
import openpyxl
from fastapi.responses import Response

power_flow_solver = PowerFlowSolver(s_base=100.0, tol=1e-4, max_iter=30)
last_simulation_result = None


@app.post("/run_simulation")
async def run_simulation(request: Request):
    global last_simulation_result
    try:
        data = await request.json()
        elements = data.get("elements", [])
        print(f"\n⚡ [조류계산 요청] {len(elements)}개의 부품 데이터를 수신하여 계산을 시작합니다.")

        result = power_flow_solver.solve(elements)
        if result.get("status") == "error":
            print(f"⚠️ 조류계산 실패: {result.get('message')}")
            return result

        last_simulation_result = result

        converged = result.get("converged", False)
        if not converged:
            print(f"⚠️ 조류계산 미수렴(발산): {result['iterations']}회 반복, 최대 오차: {result['max_mismatch']}")
            return {
                "status": "warning",
                "message": f"⚠️ 조류 계산 미수렴 (발산): {result['iterations']}회 반복 후에도 허용 오차 내에 수렴하지 못했습니다. (최대 오차: {result['max_mismatch']:.4e})",
                "data": result,
            }

        print(
            f"✅ 조류계산 성공 (수렴 완료: {result['converged']}, "
            f"{result['iterations']}회 반복, 슬랙: #{result['slack_bus']})"
        )
        print(
            f"📊 총 발전: {result['summary']['total_gen_p_mw']} MW | "
            f"총 부하: {result['summary']['total_load_p_mw']} MW | "
            f"손실: {result['summary']['total_loss_p_mw']} MW"
        )

        return {
            "status": "success",
            "message": f"조류 계산 수렴 완료! ({result['iterations']}회 반복, 최대 오차: {result['max_mismatch']})",
            "data": result,
        }
    except Exception as e:
        print(f"❌ 조류계산 중 예외 발생: {e}")
        import traceback
        traceback.print_exc()
        return {"status": "error", "message": f"조류계산 중 오류가 발생했습니다: {str(e)}"}


@app.get("/download_result_csv")
async def download_result_csv():
    if not last_simulation_result or "csv_text" not in last_simulation_result:
        return Response(content="Bus,Volt,Angle,Pgen,Qgen,Pload,Qload\n", media_type="text/csv")
    csv_bytes = last_simulation_result["csv_text"].encode("utf-8-sig")
    return Response(
        content=csv_bytes,
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="power_flow_result.csv"'},
    )


@app.get("/download_result_excel")
async def download_result_excel():
    if not last_simulation_result:
        return {"status": "error", "message": "먼저 조류계산을 실행해 주세요."}

    wb = openpyxl.Workbook()
    # Sheet 1: Bus Results
    ws_bus = wb.active
    ws_bus.title = "Bus Results"
    ws_bus.append(["Bus", "Volt (pu)", "Angle (deg)", "Pgen (MW)", "Qgen (MVAR)", "Pload (MW)", "Qload (MVAR)", "Type"])
    for r in last_simulation_result.get("bus_results", []):
        ws_bus.append([
            r.get("bus"),
            r.get("volt"),
            r.get("angle"),
            r.get("pgen"),
            r.get("qgen"),
            r.get("pload"),
            r.get("qload"),
            r.get("type"),
        ])

    # Sheet 2: Line Flows
    ws_line = wb.create_sheet(title="Line Flows")
    ws_line.append(["Line", "From Bus", "To Bus", "P From (MW)", "Q From (MVAR)", "P To (MW)", "Q To (MVAR)", "Loss P (MW)", "Loss Q (MVAR)"])
    for r in last_simulation_result.get("line_results", []):
        ws_line.append([
            r.get("label"),
            r.get("from_bus"),
            r.get("to_bus"),
            r.get("p_from_mw"),
            r.get("q_from_mvar"),
            r.get("p_to_mw"),
            r.get("q_to_mvar"),
            r.get("loss_p_mw"),
            r.get("loss_q_mvar"),
        ])

    # Sheet 3: Summary
    ws_sum = wb.create_sheet(title="Summary")
    summary = last_simulation_result.get("summary", {})
    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["Converged", last_simulation_result.get("converged")])
    ws_sum.append(["Iterations", last_simulation_result.get("iterations")])
    ws_sum.append(["Slack Bus", f"#{last_simulation_result.get('slack_bus')}"])
    ws_sum.append(["Total Gen P (MW)", summary.get("total_gen_p_mw")])
    ws_sum.append(["Total Gen Q (MVAR)", summary.get("total_gen_q_mvar")])
    ws_sum.append(["Total Load P (MW)", summary.get("total_load_p_mw")])
    ws_sum.append(["Total Load Q (MVAR)", summary.get("total_load_q_mvar")])
    ws_sum.append(["Total Loss P (MW)", summary.get("total_loss_p_mw")])
    ws_sum.append(["Total Loss Q (MVAR)", summary.get("total_loss_q_mvar")])

    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="power_flow_result.xlsx"'},
    )


@app.get("/debug_last_result")
async def debug_last_result():
    if not last_simulation_result:
        return {"msg": "no result"}
    return {
        "converged": last_simulation_result.get("converged"),
        "iterations": last_simulation_result.get("iterations"),
        "max_mismatch": last_simulation_result.get("max_mismatch"),
        "total_buses": last_simulation_result.get("total_buses"),
        "total_branches": last_simulation_result.get("total_branches"),
        "line_results": last_simulation_result.get("line_results", [])[:20],
        "summary": last_simulation_result.get("summary"),
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)
